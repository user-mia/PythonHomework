"""
第九周 Excel
"""
from openpyxl import Workbook, load_workbook

# 新建汇总表
transcript = Workbook()

# 加载成绩分析表
summary_table = load_workbook('数据分析期末成绩.xlsx')

# 获取工作簿
ws_transcript = transcript.active
ws_summary = summary_table.active

# 改名
ws_transcript.title = '期末成绩汇总测试'

# 获取成绩分析表中的所有值
ws_summary = list(ws_summary.values)


def extract(item):
    """
    提取 学号,姓名,总成绩\n
    :param item: list
    :return: list [学号,姓名,总成绩]
    """
    item = list(item)
    item[2:2] = item[7:8]
    # None 处理
    item[2] = 0 if item[2] is None else item[2]
    return item[0:3]


# 提取学号，姓名，成绩
ws_summary = [extract(item) for item in ws_summary]

# 写入标题
ws_transcript.append(ws_summary[0])

# 将前10名写入文件
for item in sorted(ws_summary[1:], key=lambda x: x[2], reverse=True)[0:10]:
    ws_transcript.append(item)

# 保存文件
transcript.save('期末成绩汇总.xlsx')
